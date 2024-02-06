import openpyxl as px
import tkinter as tk
from tkinter import messagebox

def binary_search(sheet_obj, target):
    left = 2  # Starting row index (assuming headers are in row 1)
    right = sheet_obj.max_row
    while left <= right:
        mid = (left + right) // 2
        mid_val = sheet_obj.cell(row=mid, column=1).value.lower()  # Assuming names are in the first column
        if mid_val == target:
            return mid
        elif mid_val < target:
            left = mid + 1
        else:
            right = mid - 1
    return -1  # Return -1 if not found

def validate_inputs(values):
    # Check if the input values match the specified format
    if len(values) != 5:
        return False
    if not values[0].isalpha():
        return False
    if not values[1].isdigit() or len(values[1]) != 10:
        return False
    if '@' not in values[2]:
        return False
    if not values[3].startswith('AP') or not values[3][2:].isdigit() or len(values[3]) != 13:
        return False
    return True

def add_contact(sheet_obj):
    values = get_input_values()

    if validate_inputs(values):
        new_row_index = binary_search(sheet_obj, values[0].lower())
        sheet_obj.insert_rows(new_row_index)
        for col, value in enumerate(values, start=1):
            sheet_obj.cell(row=new_row_index, column=col, value=value)
        wb_obj.save("dbms.xlsx")
        messagebox.showinfo("Success", "Contact added successfully!")
        clear_entries()
    else:
        messagebox.showerror("Error", "Invalid input format. Please check your inputs and try again.")

def search_contact(sheet_obj):
    name = get_name_entry().strip().lower()
    result_index = binary_search(sheet_obj, name)
    if result_index != -1:
        display_contact_details(sheet_obj, result_index)
    else:
        messagebox.showerror("Contact Not Found", "Contact not found in the database.")

def update_contact(sheet_obj):
    name = get_name_entry().strip().lower()
    result_index = binary_search(sheet_obj, name)
    if result_index != -1:
        create_update_window(sheet_obj, result_index)
    else:
        messagebox.showerror("Contact Not Found", "Contact not found in the database.")

def get_input_values():
    return [
        get_name_entry().strip(),
        get_mobile_entry().strip(),
        get_email_entry().strip(),
        get_reg_entry().strip(),
        get_course_entry().strip()
    ]

def get_name_entry():
    return name_entry.get()

def get_mobile_entry():
    return mobile_entry.get()

def get_email_entry():
    return email_entry.get()

def get_reg_entry():
    return reg_entry.get()

def get_course_entry():
    return course_entry.get()

def display_contact_details(sheet_obj, result_index):
    contact_details = ""
    for j in range(1, sheet_obj.max_column + 1):
        cell_obj = sheet_obj.cell(row=result_index, column=j)
        contact_details += f"{sheet_obj.cell(row=1, column=j).value}: {cell_obj.value}\n"
    messagebox.showinfo("Contact Found", contact_details)

def create_update_window(sheet_obj, result_index):
    update_window = tk.Toplevel()
    update_window.title("Update Contact")

    tk.Label(update_window, text="Select Field to Update:").pack()

    def update_field(field):
        update_window.destroy()
        create_field_update_window(sheet_obj, result_index, field)

    for field in ["Mobile Number", "Email ID", "Registration Number", "Course Enrolled"]:
        tk.Button(update_window, text=field, command=lambda f=field: update_field(f)).pack()

def create_field_update_window(sheet_obj, result_index, field):
    update_window = tk.Toplevel()
    update_window.title("Update Contact")

    tk.Label(update_window, text=f"Enter new {field}").pack()
    new_value_entry = tk.Entry(update_window)
    new_value_entry.pack()

    def save_update():
        new_value = new_value_entry.get().strip()
        sheet_obj.cell(row=result_index, column=1, value=new_value)
        wb_obj.save("dbms.xlsx")
        messagebox.showinfo("Success", f"{field} updated successfully!")
        update_window.destroy()

    tk.Button(update_window, text="Save", command=save_update).pack()

def clear_entries():
    for entry in [name_entry, mobile_entry, email_entry, reg_entry, course_entry]:
        entry.delete(0, tk.END)

def run_gui():
    global name_entry, mobile_entry, email_entry, reg_entry, course_entry, wb_obj
    wb_obj = px.load_workbook("dbms.xlsx")
    sheet_obj = wb_obj.active

    root = tk.Tk()
    root.title("Contact Management System")

    center_window(root, 500, 500)

    create_labels(root)
    create_entry_fields(root)
    create_buttons(root, sheet_obj)

    root.mainloop()

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_coordinate = (screen_width - width) / 2
    y_coordinate = (screen_height - height) / 2
    window.geometry(f"{width}x{height}+{int(x_coordinate)}+{int(y_coordinate)}")

def create_labels(root):
    labels = ["Name:", "Mobile Number:", "Email ID:", "Registration Number:", "Course Enrolled:"]
    for i, label_text in enumerate(labels):
        tk.Label(root, text=label_text).grid(row=i, column=0, padx=10, pady=5, sticky="e")

def create_entry_fields(root):
    global name_entry, mobile_entry, email_entry, reg_entry, course_entry
    entries = [name_entry, mobile_entry, email_entry, reg_entry, course_entry] = [tk.Entry(root) for _ in range(5)]
    for i, entry in enumerate(entries):
        entry.grid(row=i, column=1, padx=10, pady=5, sticky="w")

def create_buttons(root, sheet_obj):
    buttons = [
        ("Add Contact", lambda: add_contact(sheet_obj)),
        ("Search Contact", lambda: search_contact(sheet_obj)),
        ("Update Contact", lambda: update_contact(sheet_obj))
    ]
    for i, (btn_text, cmd) in enumerate(buttons, start=5):
        tk.Button(root, text=btn_text, command=cmd).grid(row=i, column=0, columnspan=2, pady=5)

if __name__ == "__main__":
    run_gui()
